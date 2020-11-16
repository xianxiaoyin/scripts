'''
Author: xianxiaoyin
LastEditors: xianxiaoyin
Descripttion: 设置打印属性
    计算表格行和列数量，如果行大于列并且大于一定行数，就把表格纵向设置，否则就横向放置
Date: 2020-10-24 10:14:59
LastEditTime: 2020-11-16 21:14:30

'''
import openpyxl
from openpyxl.worksheet.page import PrintPageSetup
import os
import sys
import zipfile
import win32com.client as win32

# 解压zip
def unzip(filename):
    _, suffix = os.path.splitext(filename)
    if suffix == ".zip":
        with zipfile.ZipFile(filename) as zf:
            path, _ = os.path.split(filename)
            zf.extractall(path)

# 处理中文乱码的文件名


def changeFilename(path):
    for filename in os.listdir(path):
        try:
            zip_file = filename.encode('cp437').decode('gbk')
        except:
            zip_file = filename.encode('utf-8').decode('utf-8')
        if os.path.isdir(os.path.join(path, filename)):
            os.rename(os.path.join(path, filename),
                      os.path.join(path, zip_file))
            changeFilename(os.path.join(path, zip_file))
        else:
            os.rename(os.path.join(path, filename),
                      os.path.join(path, zip_file))


def get_files(path, file_list=[]):
    for file in os.listdir(path):
        if os.path.isdir(os.path.join(path, file)):
            get_files(os.path.join(path, file))
        else:
            _, suffix = os.path.splitext(os.path.join(path, file))
            if suffix == ".xls":
                file_list.append(os.path.join(path, file))

    return file_list

# 转换文件格式
def xlsToxlsx(filename, formats):
    if formats == 51:  # .xls-->.xlsx
        new_filename = "{}x".format(filename)
    elif formats == 56:  # .xlsx-->.xls
        pass
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(filename)
    wb.SaveAs(new_filename, FileFormat=formats)
    wb.Close()
    excel.Application.Quit()

def MaxRows(ws, line=50):
    if ws.max_row > ws.max_column and ws.max_row > line:
        return True


def formatPage(fileName, newFileName):
    wb = openpyxl.load_workbook(fileName)
    for index, _ in enumerate(wb.sheetnames):
        ws = wb[wb.sheetnames[index]]
        # 设置页面
        if MaxRows(ws):
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT     # 页面纵向
        else:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE    # 页面横向
        ws.page_setup.fitToHeight = True                            # 自适应高度
        ws.sheet_properties.pageSetUpPr.fitToPage = True            # 自适应页面宽度（将表格放在一页）
        ws.print_options.horizontalCentered = True                  # 水平居中
        ws.print_options.verticalCentered = True                    # 垂直居中
    wb.save(newFileName)


def main():
    for file_zip in os.listdir(os.getcwd()):
        filename_zip = os.path.join(os.getcwd(), file_zip)
        unzip(filename_zip)
    changeFilename(os.getcwd())
    for filename in get_files(os.getcwd()):
        xlsToxlsx(filename, 51) 
        xlsxFileName = "{}x".format(filename)
        path, fname = os.path.split(xlsxFileName)
        tmpPath = path.split(os.sep)
        index =  len(tmpPath) - len(os.getcwd().split(os.sep))
        tmpPath.insert(-index, "new")
        newPath = str(os.sep).join(tmpPath)
        if not os.path.exists(newPath):
            os.makedirs(newPath)
        newXlsxFfileName = os.path.join(newPath, fname)
        formatPage(xlsxFileName, newXlsxFfileName)


if __name__ == '__main__':
    main()