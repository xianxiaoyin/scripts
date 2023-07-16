"""
Author: xianxiaoyin
LastEditors: xianxiaoyin
Descripttion: 先将xls文件转成xlsx的文件，然后格式化数据，然后再将xlsx文件转成xls [设置行高，宽度之类]
Date: 2020-11-09 22:30:26
LastEditTime: 2020-11-16 21:13:27
"""
import os
import sys
import zipfile

import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


# 解压zip
def unzip(filename):
    _, suffix = os.path.splitext(filename)
    if suffix == ".zip":
        with zipfile.ZipFile(filename) as zf:
            path, _ = os.path.split(filename)
            zf.extractall(path)


# 处理中文乱码的文件名


def changeFilename(path):
    for filename in os.listdir(path):
        # if os.path.isdir(os.path.join(path, filename)):
        try:
            zip_file = filename.encode("cp437").decode("gbk")
        except:
            zip_file = filename.encode("utf-8").decode("utf-8")
        if os.path.isdir(os.path.join(path, filename)):
            os.rename(os.path.join(path, filename), os.path.join(path, zip_file))
            changeFilename(os.path.join(path, zip_file))
        else:
            os.rename(os.path.join(path, filename), os.path.join(path, zip_file))


def get_files(path, file_list=[]):
    for file in os.listdir(path):
        if os.path.isdir(os.path.join(path, file)):
            get_files(os.path.join(path, file))
        else:
            _, suffix = os.path.splitext(os.path.join(path, file))
            if suffix == ".xls":
                file_list.append(os.path.join(path, file))

    return file_list


# 转换文件格式
def xlsToxlsx(filename, formats):
    if formats == 51:  # .xls-->.xlsx
        new_filename = "{}x".format(filename)
    elif formats == 56:  # .xlsx-->.xls
        pass
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(filename)
    wb.SaveAs(new_filename, FileFormat=formats)
    wb.Close()
    excel.Application.Quit()


# 检查所有的单元格，如果存在时间的单元格记录对应的列
def checkDataFormat(ws):
    cell_list = []
    row_list = []

    tag = [
        "yyyy/m/d;@",
        "[$-F800]dddd\,\ mmmm\ dd\,\ yyyy",
        "mm-dd-yy",
        'm"月"d"日";@',
        'yyyy"年"m"月"d"日";@',
    ]
    for row in ws.rows:
        for cell in row:
            # print(cell.value)
            # print(cell.number_format)
            # if cell.value and cell.value == "1234567890":
            if cell.value and int(cell.font.sz) > 13:
                row_list.append(cell.row)
            if cell.value:
                if cell.number_format in tag:
                    # if cell.value and cell.number_format != "General":
                    cell_list.append(get_column_letter(cell.column))
    return set(cell_list), set(row_list)


# 自动给excel表格修改行高
# 1.xlsx 的文件修改完成之后变成1_new.xlsx
def formatExcel(filename, filename2, height=13, width=15):
    wb = load_workbook(filename)
    for index, _ in enumerate(wb.sheetnames):
        ws = wb[wb.sheetnames[index]]

        cdf, cdf2 = checkDataFormat(ws)
        for i in range(1, ws.max_row + 1):
            if i in cdf2:
                ws.row_dimensions[i].height = height + 4
            else:
                ws.row_dimensions[i].height = height

        for j in range(1, ws.max_column + 1):
            if get_column_letter(j) in cdf:
                ws.column_dimensions[get_column_letter(j)].width = width
            else:
                ws.column_dimensions[get_column_letter(j)].width = width - 3

    wb.save(filename2)


def main():
    for file_zip in os.listdir(os.getcwd()):
        filename_zip = os.path.join(os.getcwd(), file_zip)
        unzip(filename_zip)
    changeFilename(os.getcwd())
    for filename in get_files(os.getcwd()):
        xlsToxlsx(filename, 51)
        xlsxFileName = f"{filename}x"
        path, fname = os.path.split(xlsxFileName)
        tmpPath = path.split(os.sep)
        index = len(tmpPath) - len(os.getcwd().split(os.sep))
        tmpPath.insert(-index, "new")
        newPath = str(os.sep).join(tmpPath)
        if not os.path.exists(newPath):
            os.makedirs(newPath)
        newXlsxFfileName = os.path.join(newPath, fname)
        formatExcel(xlsxFileName, newXlsxFfileName, 12)


if __name__ == "__main__":
    main()
